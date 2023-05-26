import openpyxl

def main():
    workbook = openpyxl.Workbook()
    #ws = workbook.worksheets[0]
    ws = workbook.active
    cell = ws.cell(row=1, column=1)
    cell.value = "Hello World"
    cell.font = openpyxl.styles.Font(color="EE5555", bold=True)
    cell.alignment = openpyxl.styles.Alignment(horizontal="center")
    #cell.fill = openpyxl.styles.PatternFill(bgColor="FFC7CE", fill_type="solid")  # do sprawdzenia

    ws["B2"].value = 5
    ws["B2"].comment = openpyxl.comments.Comment("To jest komentarz", "kurs")
    ws["B3"].value = "link"
    ws["B3"].hyperlink = "https://www.google.com"

    ws.column_dimensions["A"].width = 50
    assert openpyxl.utils.get_column_letter(1) == "A"

    workbook.save("test.xlsx")
    workbook.close()

def load_example():
    wb = openpyxl.load_workbook(r"C:\Users\kurs\Downloads\dane\excel_example.xlsx")
    ws = wb.active
    for row in ws.iter_rows(min_row=2):
        print(row[0].value, row[4].value)  # tu wyjątkowo w openpyxl indeksowanie od zera!


if __name__ == "__main__":
    load_example()
    # main()